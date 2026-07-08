"""
Estimate generation and retrieval routes for BuildSmart AI Estimator.
Calls the estimator engine and persists results (estimate + BOQ items) to Supabase.
"""

import io
from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from services.db import supabase_admin, get_supabase_user
from services.estimator import calculate_estimate
from services.pdf_generator import generate_estimate_pdf

estimates_bp = Blueprint('estimates_bp', __name__, url_prefix='/api/estimates')


# ---------------------------------------------------------------------------
# Helpers for database check constraint compatibility
# Maps 'Base' (UI) to 'Budget' (DB constraint)
# ---------------------------------------------------------------------------
def map_quality_to_db(q):
    if q == 'Base':
        return 'Budget'
    return q

def map_quality_from_db(q):
    if q == 'Budget':
        return 'Base'
    return q




# ---------------------------------------------------------------------------
# Helper – reuse same token extractor as projects
# ---------------------------------------------------------------------------
def get_user_id_from_token(req):
    """
    Extract and validate Bearer token, returning the user UUID.

    Raises:
        ValueError: on missing / invalid token.
    """
    auth_header = req.headers.get('Authorization', '')
    if not auth_header.startswith('Bearer '):
        raise ValueError('Missing or invalid Authorization header.')

    token = auth_header[len('Bearer '):]

    try:
        user_response = get_supabase_user(token)
        user = user_response.user
    except Exception as exc:
        raise ValueError(f'Token validation failed: {exc}')

    if not user:
        raise ValueError('Invalid or expired token.')

    return user.id


# ---------------------------------------------------------------------------
# POST /api/estimates/generate
# ---------------------------------------------------------------------------
@estimates_bp.route('/generate', methods=['POST'])
def generate_estimate():
    """
    Generate a cost estimate using the estimator engine and persist everything.

    Body (all estimate inputs + optional project_id):
        project_id       – (optional) existing project UUID to link estimate to
        name             – project name (used when auto-creating a project)
        location / city  – city for rate lookup
        building_type    – Residential / Commercial / etc.
        total_sqft       – built-up area in sq ft
        floors           – number of floors
        bedrooms         – number of bedrooms
        bathrooms        – number of bathrooms
        quality          – Standard / Premium / Luxury
        ... (any additional params accepted by calculate_estimate)
    """
    try:
        user_id = get_user_id_from_token(request)
    except ValueError as ve:
        return jsonify({'error': str(ve)}), 401

    try:
        data    = request.get_json(force=True) or {}

        # Validate customer email and phone if provided
        import re
        cust_email = data.get('customer_email') or data.get('customerEmail')
        if cust_email and str(cust_email).strip():
            email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
            if not re.match(email_regex, str(cust_email).strip()):
                return jsonify({'error': 'Please enter a valid customer email address.'}), 400

        cust_mobile = data.get('customer_mobile') or data.get('customerMobile')
        if cust_mobile and str(cust_mobile).strip():
            phone_regex = r'^[6-9]\d{9}$'
            if not re.match(phone_regex, str(cust_mobile).strip()):
                return jsonify({'error': 'Please enter a valid 10-digit customer mobile number starting with 6-9.'}), 400

        # ── Resolve or create project ──────────────────────────────────────
        project_id = data.get('project_id')

        if project_id:
            # Verify the project belongs to this user
            ownership = (
                supabase_admin
                .table('projects')
                .select('id')
                .eq('id', project_id)
                .eq('builder_id', user_id)
                .single()
                .execute()
            )
            if not ownership.data:
                return jsonify({'error': 'Project not found or access denied.'}), 404
        else:
            # Auto-create a project from the incoming params
            project_name = (
                data.get('name')
                or data.get('project_name')
                or 'Untitled Project'
            )
            
            project_payload = {
                'builder_id':    user_id,
                'name':          project_name,
                'location':      data.get('location') or data.get('city', ''),
                'building_type': data.get('building_type') or data.get('buildingType') or 'Residential Villa',
                'total_sqft':    data.get('total_sqft') or data.get('totalSqft') or 1500,
                'floors':        data.get('floors') or 1,
                'bedrooms':      data.get('bedrooms') or 0,
                'bathrooms':     data.get('bathrooms') or 0,
                'quality':       map_quality_to_db(data.get('quality', 'Standard')),
            }
            
            # Map new optional columns
            for col in [
                'customer_name', 'customer_mobile', 'customer_email', 'project_name',
                'quotation_number', 'quotation_date', 'builder_company_name',
                'country', 'state', 'city', 'locality', 'pincode', 'project_address',
                'structure_type', 'plot_area_sqft', 'builtup_area_sqft', 'basement_area_sqft',
                'parking_area_sqft', 'terrace_area_sqft', 'balcony_area_sqft',
                'total_construction_area_sqft', 'living_rooms', 'dining_rooms', 'kitchens',
                'pooja_rooms', 'study_rooms', 'store_rooms', 'soil_type',
                'contingency_percentage', 'builder_margin_percentage', 'gst_percentage'
            ]:
                val = data.get(col)
                if val is not None and val != "":
                    # handle types
                    if col in ['plot_area_sqft', 'builtup_area_sqft', 'basement_area_sqft', 'parking_area_sqft', 'terrace_area_sqft', 'balcony_area_sqft', 'total_construction_area_sqft', 'contingency_percentage', 'builder_margin_percentage', 'gst_percentage']:
                        project_payload[col] = float(val)
                    elif col in ['living_rooms', 'dining_rooms', 'kitchens', 'pooja_rooms', 'study_rooms', 'store_rooms', 'floors', 'bedrooms', 'bathrooms']:
                        project_payload[col] = int(val)
                    else:
                        project_payload[col] = str(val)

            new_project = (
                supabase_admin
                .table('projects')
                .insert(project_payload)
                .execute()
            )
            project_id = new_project.data[0]['id']

            # Dynamic floors insert
            floors_list = data.get('floors_list') or data.get('floorsList') or []
            if floors_list:
                floors_payload = [
                    {
                        'project_id': project_id,
                        'floor_name': f.get('floor_name') or f.get('name') or 'Floor',
                        'floor_area_sqft': float(f.get('floor_area_sqft') or f.get('area') or 0)
                    }
                    for f in floors_list
                ]
                supabase_admin.table('project_floors').insert(floors_payload).execute()

            # Dynamic flooring rooms insert
            flooring_rooms = data.get('flooring_rooms') or []
            if flooring_rooms:
                rooms_payload = [
                    {
                        'project_id': project_id,
                        'room_name': r.get('room_name') or r.get('name') or 'Room',
                        'flooring_type': r.get('flooring_type') or r.get('type') or 'vitrified',
                        'flooring_brand': r.get('flooring_brand') or r.get('brand') or '',
                        'flooring_quality': r.get('flooring_quality') or r.get('quality') or '',
                        'area_sqft': float(r.get('area_sqft') or r.get('area') or 0)
                    }
                    for r in flooring_rooms
                ]
                supabase_admin.table('project_rooms_flooring').insert(rooms_payload).execute()


        # ── Run the estimation engine ──────────────────────────────────────
        result = calculate_estimate(data)

        costs    = result.get('costs', {})
        duration = result.get('duration', {})

        # ── Determine next version number for this project ─────────────────
        existing_versions = (
            supabase_admin
            .table('estimates')
            .select('version')
            .eq('project_id', project_id)
            .order('version', desc=True)
            .limit(1)
            .execute()
        )
        last_version = (
            existing_versions.data[0]['version']
            if existing_versions.data
            else 0
        )
        next_version = last_version + 1

        # ── Persist the estimate header ────────────────────────────────────
        estimate_insert = (
            supabase_admin
            .table('estimates')
            .insert({
                'project_id':        project_id,
                'version':           next_version,
                'input_json':        data,
                'output_json': {
                    'recommendations': result.get('recommendations', []),
                    'quantities':      result.get('quantities', {}),
                    'duration':        duration,
                    'stages_breakdown': result.get('stages_breakdown', []),
                },
                'subtotal':          costs.get('subtotal', 0),
                'contingency_pct':   costs.get('contingencyPct', 0),
                'contingency_amount': costs.get('contingency', 0),
                'gst_pct':           costs.get('gstPct', 0),
                'gst_amount':        costs.get('gst', 0),
                'grand_total':       costs.get('grandTotal', 0),
                'duration_min':      duration.get('min', 0),
                'duration_max':      duration.get('max', 0),
                'status':            'Draft',
            })
            .execute()
        )

        estimate = estimate_insert.data[0]
        estimate_id = estimate['id']

        # ── Persist BOQ line items ─────────────────────────────────────────
        boq_items = result.get('boqItems', result.get('boq', []))
        if boq_items:
            items_payload = [
                {
                    'estimate_id':     estimate_id,
                    'category':        item.get('category', ''),
                    'item_code':       item.get('code', ''),
                    'description':     item.get('description', ''),
                    'qty':             item.get('qty', 0),
                    'unit':            item.get('unit', ''),
                    'rate':            item.get('rate', 0),
                    'amount':          item.get('amount', 0),
                }
                for item in boq_items
            ]
            supabase_admin.table('estimate_items').insert(items_payload).execute()


        # ── Return full result ─────────────────────────────────────────────
        return jsonify({
            'success':         True,
            'estimate_id':     estimate_id,
            'project_id':      project_id,
            'version':         next_version,
            'costs':           result.get('costs', {}),
            'duration':        result.get('duration', {}),
            'boqItems':        boq_items,
            'quantities':      result.get('quantities', {}),
            'recommendations': result.get('recommendations', []),
            'stages_breakdown': result.get('stages_breakdown', []),
            'summary':         result.get('summary', {}),
            'input':           result.get('input', data),
        }), 201

    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(exc)}), 500


# ---------------------------------------------------------------------------
# GET /api/estimates/project/<project_id>
# ---------------------------------------------------------------------------
@estimates_bp.route('/project/<project_id>', methods=['GET'])
def list_estimates_for_project(project_id):
    """
    Return all estimates for a given project (ordered newest first).
    """
    try:
        get_user_id_from_token(request)   # just authenticate; no ownership filter here

        response = (
            supabase_admin
            .table('estimates')
            .select('*')
            .eq('project_id', project_id)
            .order('version', desc=True)
            .execute()
        )

        return jsonify({'success': True, 'data': response.data}), 200

    except ValueError as ve:
        return jsonify({'error': str(ve)}), 401
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


# ---------------------------------------------------------------------------
# GET /api/estimates/<estimate_id>
# ---------------------------------------------------------------------------
@estimates_bp.route('/<estimate_id>', methods=['GET'])
def get_estimate(estimate_id):
    """
    Return a single estimate with all its BOQ line items.
    """
    try:
        get_user_id_from_token(request)

        # Fetch estimate header
        estimate_resp = (
            supabase_admin
            .table('estimates')
            .select('*')
            .eq('id', estimate_id)
            .single()
            .execute()
        )

        if not estimate_resp.data:
            return jsonify({'error': 'Estimate not found.'}), 404

        estimate = estimate_resp.data

        # Fetch associated BOQ items
        items_resp = (
            supabase_admin
            .table('estimate_items')
            .select('*')
            .eq('estimate_id', estimate_id)
            .order('category')
            .execute()
        )

        estimate['items'] = items_resp.data or []

        # Fetch project details including floors and room flooring
        project_id = estimate.get('project_id')
        if project_id:
            try:
                proj_resp = supabase_admin.table('projects').select('*').eq('id', project_id).single().execute()
                estimate['project'] = proj_resp.data or {}
                if estimate['project'] and 'quality' in estimate['project']:
                    estimate['project']['quality'] = map_quality_from_db(estimate['project']['quality'])
                
                floors_resp = supabase_admin.table('project_floors').select('*').eq('project_id', project_id).execute()
                estimate['project']['floors_list'] = floors_resp.data or []
                
                rooms_resp = supabase_admin.table('project_rooms_flooring').select('*').eq('project_id', project_id).execute()
                estimate['project']['flooring_rooms'] = rooms_resp.data or []
            except Exception as e:
                print(f"Error fetching child project details: {e}")

        return jsonify({'success': True, 'data': estimate}), 200

    except ValueError as ve:
        return jsonify({'error': str(ve)}), 401
    except Exception as exc:
        if 'No rows found' in str(exc) or 'PGRST116' in str(exc):
            return jsonify({'error': 'Estimate not found.'}), 404
        return jsonify({'error': str(exc)}), 500


# ---------------------------------------------------------------------------
# GET /api/estimates/<estimate_id>/export/excel
# ---------------------------------------------------------------------------
@estimates_bp.route('/<estimate_id>/export/excel', methods=['GET'])
def export_estimate_excel(estimate_id):
    """
    Generate and download a beautifully formatted Excel sheet containing
    the estimate summary and full itemized BOQ.
    """
    try:
        get_user_id_from_token(request)

        # 1. Fetch estimate header
        estimate_resp = (
            supabase_admin
            .table('estimates')
            .select('*, projects(*)')
            .eq('id', estimate_id)
            .single()
            .execute()
        )

        if not estimate_resp.data:
            return jsonify({'error': 'Estimate not found.'}), 404

        estimate = estimate_resp.data
        project = estimate.get('projects', {})
        if project and 'quality' in project:
            project['quality'] = map_quality_from_db(project['quality'])
        output_json = estimate.get('output_json', {})

        # 2. Fetch BOQ items
        items_resp = (
            supabase_admin
            .table('estimate_items')
            .select('*')
            .eq('estimate_id', estimate_id)
            .order('category')
            .execute()
        )
        boq_items = items_resp.data or []

        # 3. Build Excel Workbook
        wb = Workbook()

        # Sheet 1: Summary
        ws1 = wb.active
        ws1.title = "Estimate Summary"
        ws1.views.sheetView[0].showGridLines = True

        # Styles
        title_font = Font(name='Arial', size=16, bold=True, color='0F766E')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        section_font = Font(name='Arial', size=12, bold=True, color='1E293B')
        bold_font = Font(name='Arial', size=10, bold=True)
        regular_font = Font(name='Arial', size=10)
        
        header_fill = PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid')
        accent_fill = PatternFill(start_color='CCFBF1', end_color='CCFBF1', fill_type='solid')
        gray_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        double_bottom = Border(bottom=Side(style='double', color='0F766E'), top=Side(style='thin', color='CBD5E1'))

        # Title block
        ws1['A1'] = "Buildsmart 360 - Project Quotation"
        ws1['A1'].font = title_font
        ws1.row_dimensions[1].height = 25

        # Project Info Section
        ws1['A3'] = "Project Information"
        ws1['A3'].font = section_font
        
        info_labels = [
            ("Project Name:", project.get('project_name') or project.get('name', 'Untitled')),
            ("Customer Name:", project.get('customer_name') or '—'),
            ("Quotation Number:", project.get('quotation_number') or '—'),
            ("Location/City:", project.get('city') or project.get('location', 'Bangalore')),
            ("Building Type:", project.get('building_type', 'Residential')),
            ("Built-up Area:", f"{project.get('total_sqft', 0):,} Sqft"),
            ("Structure Type:", project.get('structure_type') or 'RCC Frame'),
            ("Soil Type:", project.get('soil_type') or 'Normal Soil'),
            ("Quality Grade:", project.get('quality', 'Standard')),
            ("Version:", f"v{estimate.get('version', 1)}"),
            ("Generated On:", (estimate.get('generated_at') or '').split('T')[0])
        ]

        
        row_idx = 4
        for label, val in info_labels:
            ws1.cell(row=row_idx, column=1, value=label).font = bold_font
            ws1.cell(row=row_idx, column=2, value=val).font = regular_font
            row_idx += 1

        # Cost Breakdown Section
        row_idx += 1
        ws1.cell(row=row_idx, column=1, value="Cost Breakdown").font = section_font
        
        row_idx += 1
        cost_headers = ["Component / Cost Category", "Amount (INR)"]
        for col_idx, h in enumerate(cost_headers, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='left')

        costs = [
            ("Grand Total Estimate", estimate.get('grand_total', 0))
        ]

        for label, val in costs:
            row_idx += 1
            l_cell = ws1.cell(row=row_idx, column=1, value=label)
            v_cell = ws1.cell(row=row_idx, column=2, value=val)
            
            l_cell.font = bold_font if label == "Grand Total Estimate" else regular_font
            v_cell.font = bold_font if label == "Grand Total Estimate" else regular_font
            v_cell.number_format = '₹#,##,##0'
            
            if label == "Grand Total Estimate":
                l_cell.fill = accent_fill
                v_cell.fill = accent_fill
                l_cell.border = double_bottom
                v_cell.border = double_bottom
            else:
                l_cell.border = thin_border
                v_cell.border = thin_border

        # AI Recommendations Section
        row_idx += 2
        ws1.cell(row=row_idx, column=1, value="AI-Powered Recommendations & Insights").font = section_font
        
        recs = output_json.get('recommendations', [])
        if recs:
            for rec in recs:
                row_idx += 1
                rec_title = f"{rec.get('type', 'Tip').upper()}: {rec.get('title', '')}"
                ws1.cell(row=row_idx, column=1, value=rec_title).font = bold_font
                
                row_idx += 1
                ws1.cell(row=row_idx, column=1, value=rec.get('text', '')).font = regular_font
                ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
        else:
            row_idx += 1
            ws1.cell(row=row_idx, column=1, value="No active recommendations for this profile.").font = regular_font

        # Column widths Summary
        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 25

        # Sheet 2: BOQ
        ws2 = wb.create_sheet(title="Bill of Quantities")
        ws2.views.sheetView[0].showGridLines = True

        # Table headers
        ws2['A1'] = "Bill of Quantities (BOQ)"
        ws2['A1'].font = title_font
        
        ws2['A3'] = "Code"
        ws2['B3'] = "Category"
        ws2['C3'] = "Description"
        ws2['D3'] = "Unit"
        ws2['E3'] = "Quantity"

        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = ws2[f"{col}3"]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='left')

        # Group by category
        from collections import defaultdict
        grouped = defaultdict(list)
        for item in boq_items:
            grouped[item.get('category', 'Other')].append(item)

        row_idx = 4
        for category, items in grouped.items():
            # Category subheader row
            ws2.cell(row=row_idx, column=1, value=category).font = section_font
            ws2.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
            for c in range(1, 6):
                ws2.cell(row=row_idx, column=c).fill = gray_fill
                ws2.cell(row=row_idx, column=c).border = thin_border
            row_idx += 1

            for item in items:
                ws2.cell(row=row_idx, column=1, value=item.get('material_code', '')).font = regular_font
                ws2.cell(row=row_idx, column=2, value=category).font = regular_font
                ws2.cell(row=row_idx, column=3, value=item.get('description', '')).font = regular_font
                ws2.cell(row=row_idx, column=4, value=item.get('unit', '')).font = regular_font
                
                qty_cell = ws2.cell(row=row_idx, column=5, value=item.get('quantity', 0))
                qty_cell.font = regular_font
                qty_cell.number_format = '#,##0'

                for c in range(1, 6):
                    ws2.cell(row=row_idx, column=c).border = thin_border
                row_idx += 1

        # Grand Total row
        row_idx += 1
        total_label_cell = ws2.cell(row=row_idx, column=1, value="GRAND TOTAL ESTIMATE")
        total_label_cell.font = bold_font
        ws2.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
        
        total_amt_cell = ws2.cell(row=row_idx, column=5, value=estimate.get('grand_total', 0))
        total_amt_cell.font = bold_font
        total_amt_cell.number_format = '₹#,##0'
        
        for c in range(1, 6):
            ws2.cell(row=row_idx, column=c).fill = accent_fill
            ws2.cell(row=row_idx, column=c).border = double_bottom

        # Column widths BOQ
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 50
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 15

        # Save workbook to memory
        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)

        filename = f"BOQ_Estimate_{project.get('name', 'Project')}_v{estimate.get('version', 1)}.xlsx"
        # Sanitize filename
        import re
        filename = re.sub(r'[\\/*?:"<>|]', "", filename).replace(" ", "_")

        return send_file(
            fp,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )

    except ValueError as ve:
        return jsonify({'error': str(ve)}), 401
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


# ---------------------------------------------------------------------------
# POST /api/estimates/<estimate_id>/share-pdf  –  email estimate PDF to client & builder
# ---------------------------------------------------------------------------
@estimates_bp.route('/<estimate_id>/share-pdf', methods=['POST'])
def share_estimate_pdf(estimate_id):
    """
    Generate estimate PDF and email it to the builder and client.
    Requires: Authorization: Bearer <token>
    """
    try:
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Bearer '):
            return jsonify({'error': 'Missing or invalid Authorization header.'}), 401

        token = auth_header[len('Bearer '):]
        user_response = get_supabase_user(token)
        user = user_response.user

        if not user:
            return jsonify({'error': 'Invalid or expired token.'}), 401

        # 1. Fetch estimate header
        est_res = (
            supabase_admin
            .table('estimates')
            .select('*')
            .eq('id', estimate_id)
            .single()
            .execute()
        )
        if not est_res.data:
            return jsonify({'error': 'Estimate not found.'}), 404
        estimate = est_res.data

        # Verify builder ownership of project
        project_res = (
            supabase_admin
            .table('projects')
            .select('*')
            .eq('id', estimate['project_id'])
            .single()
            .execute()
        )
        if not project_res.data or project_res.data['builder_id'] != user.id:
            return jsonify({'error': 'Access denied.'}), 403
        project = project_res.data

        # 2. Fetch estimate items
        items_res = (
            supabase_admin
            .table('estimate_items')
            .select('*')
            .eq('estimate_id', estimate_id)
            .order('sort_order', desc=False)
            .execute()
        )
        items = items_res.data or []

        # Fetch builder profile (company_name and avatar_url)
        profile_res = (
            supabase_admin
            .table('profiles')
            .select('company_name, avatar_url')
            .eq('id', user.id)
            .single()
            .execute()
        )
        builder_profile = profile_res.data or {}
        builder_name = builder_profile.get('company_name')
        builder_logo = builder_profile.get('avatar_url')

        # 3. Generate PDF
        pdf_bytes = generate_estimate_pdf(
            estimate, items,
            builder_name=builder_name,
            builder_logo=builder_logo
        )

        # Base64 encode the PDF
        import base64
        pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')

        # 4. Determine emails
        builder_email = user.email
        inj = estimate.get('input_json') or {}
        client_email = inj.get('customer_email') or inj.get('customerEmail')
        
        if not client_email:
            return jsonify({'error': 'Client email is missing from estimate details.'}), 400

        # Send Email using Brevo REST API
        import os
        import requests
        brevo_key = os.getenv('BREVO_API_KEY')
        sender_email = os.getenv('BREVO_SENDER_EMAIL')
        sender_name = os.getenv('BREVO_SENDER_NAME', 'Vtab square')

        if not brevo_key or not sender_email:
            return jsonify({'error': 'Email gateway is not configured on the server.'}), 500

        # Formulate HTML Email body
        project_name = project.get('name', 'Residential project')
        html_content = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #e2e8f0; border-radius: 8px;">
                <h2 style="color: #0f766e; text-align: center;">Buildsmart 360</h2>
                <hr style="border: 0; border-top: 1px solid #e2e8f0; margin: 20px 0;">
                <p>Hello,</p>
                <p>Please find attached the detailed construction cost estimate and Bill of Quantities (BOQ) for the project <b>{project_name}</b>.</p>
                <p><b>Estimate Details:</b></p>
                <ul>
                    <li><b>Project Name:</b> {project_name}</li>
                    <li><b>Location:</b> {project.get('location', 'India')}</li>
                    <li><b>Total Built-up Area:</b> {project.get('total_sqft', '—')} Sqft</li>
                    <li><b>Total Estimated Price:</b> Rs. {float(estimate.get('grand_total', 0)):,.2f}</li>
                </ul>
                <p>The detailed specifications and itemized cost breakdown are attached to this email as a PDF document.</p>
                <p style="color: #64748b; font-size: 13px;">If you have any questions or require modifications, please contact the builder directly.</p>
                <hr style="border: 0; border-top: 1px solid #e2e8f0; margin: 20px 0;">
                <p style="font-size: 12px; color: #94a3b8; text-align: center;">© 2026 Buildsmart 360. Built for Indian Builders.</p>
            </div>
        </body>
        </html>
        """

        payload = {
            "sender": {"name": sender_name, "email": sender_email},
            "to": [
                {"email": client_email, "name": (inj.get('customer_name') or 'Client')},
                {"email": builder_email, "name": (user.email.split('@')[0] or 'Builder')}
            ],
            "subject": f"Construction Cost Estimate - {project_name}",
            "htmlContent": html_content,
            "attachment": [
                {
                    "content": pdf_base64,
                    "name": f"Estimate_{project_name.replace(' ', '_')}.pdf"
                }
            ]
        }

        headers = {
            "api-key": brevo_key,
            "content-type": "application/json",
            "accept": "application/json"
        }

        response = requests.post("https://api.brevo.com/v3/smtp/email", json=payload, headers=headers)
        if response.status_code not in (200, 201, 202):
            return jsonify({'error': f'Failed to send email. Gateway returned code {response.status_code}'}), 502

        return jsonify({'success': True, 'message': f'Estimate PDF successfully emailed to {client_email} and {builder_email}.'}), 200

    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


# GET /api/estimates/<estimate_id>/download-pdf  –  download estimate PDF file
# ---------------------------------------------------------------------------
@estimates_bp.route('/<estimate_id>/download-pdf', methods=['GET'])
def download_estimate_pdf(estimate_id):
    """
    Generate estimate PDF and send it as a downloadable response.
    Requires: Authorization: Bearer <token>
    """
    try:
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Bearer '):
            return jsonify({'error': 'Missing or invalid Authorization header.'}), 401
        
        token = auth_header[len('Bearer '):]
        user_response = get_supabase_user(token)
        user = user_response.user
        if not user:
            return jsonify({'error': 'Invalid or expired token.'}), 401

        # Fetch estimate header
        est_res = (
            supabase_admin
            .table('estimates')
            .select('*')
            .eq('id', estimate_id)
            .single()
            .execute()
        )
        if not est_res.data:
            return jsonify({'error': 'Estimate not found.'}), 404
        estimate = est_res.data

        # Verify builder ownership
        project_res = (
            supabase_admin
            .table('projects')
            .select('*')
            .eq('id', estimate['project_id'])
            .single()
            .execute()
        )
        if not project_res.data or project_res.data['builder_id'] != user.id:
            return jsonify({'error': 'Access denied.'}), 403
        project = project_res.data

        # Fetch estimate items
        items_res = (
            supabase_admin
            .table('estimate_items')
            .select('*')
            .eq('estimate_id', estimate_id)
            .order('sort_order', desc=False)
            .execute()
        )
        items = items_res.data or []

        # Fetch builder profile (company_name and avatar_url)
        profile_res = (
            supabase_admin
            .table('profiles')
            .select('company_name, avatar_url')
            .eq('id', user.id)
            .single()
            .execute()
        )
        builder_profile = profile_res.data or {}
        builder_name = builder_profile.get('company_name')
        builder_logo = builder_profile.get('avatar_url')

        # Generate PDF
        pdf_bytes = generate_estimate_pdf(
            estimate, items,
            builder_name=builder_name,
            builder_logo=builder_logo
        )

        project_name = project.get('name', 'Quotation')
        safe_filename = f"Estimate_{project_name.replace(' ', '_')}.pdf"
        
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=safe_filename
        )
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


